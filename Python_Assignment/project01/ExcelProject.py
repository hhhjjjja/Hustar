import tkinter
from tkinter import *
import tkinter.messagebox

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side


window=tkinter.Tk()
window.title("Excel")

thin_border = Border(left=Side(style='thin', color="000000"), right=Side(style='thin', color="000000"), top=Side(style='thin', color="000000"), bottom=Side(style='thin', color="000000"))
fill_back = openpyxl.styles.fills.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

flag = 0
compareValue=[]
all_values = []

path = "./"
filename = "[첨부2]기관별간접비 고시기준.xlsx"

def OpenEvent():
    try:
        global wb
        global ws

        # data_only=Ture -> 수식이 아닌 값으로 받아옴
        wb = load_workbook(path + filename, data_only=True)
        ws = wb.active

        #엑셀 값 넣기
        for row in ws.rows:
            row_value = []
            for cell in row:
                if (cell.value != None):
                    row_value.append(cell.value)
            all_values.append(row_value)
        #타이틀, 헤더 삭제/ 데이터만 남김
        del all_values[0]
        del all_values[0]

        print('\nOpen [' + filename + ']')

    except:
        print('\n!!!Not Found!!!')
        tkinter.messagebox.showerror("ERROR", "Not found")


def ViewEvent() :
    try:
        print('\n------------View Sheet------------')
        print(ws['A1'].value, "\n")
        print('['+ws['A2'].value+',\t'+ws['B2'].value+',\t'+ws['C2'].value+']')
        #print(all_values)
        for i,a in enumerate(all_values):
            print("[", a[0], ", ", float(a[1]), ", ", float(a[2]), "]")

    except:
        print('\n!!! Open Excel First !!!')
        tkinter.messagebox.showwarning("WARNING", "Open Excel file First")

def CompareEvent() :
    try:
        global flag
        global compareValue

        if flag == 0:
            compareValue = sorted(all_values, key=lambda x: x[1], reverse=False)
            flag = 1
            print('\n------------오름차순 정렬------------')
        elif flag == 1:
            compareValue = sorted(all_values, key=lambda x: x[1], reverse=True)
            flag = 0
            print('\n------------내림차순 정렬------------')

        for i, b in enumerate(compareValue):
            print(b[0], float(b[1]), float(b[2]))
    except:
        print('\n!!! Open Excel First !!!')
        tkinter.messagebox.showwarning("WARNING", "Open Excel file First")

def SaveEvent() :
    try:
        # 저장 팝업
        ok = tkinter.messagebox.askokcancel("Save", "저장하시겠습니까?")
        # 취소
        if not ok:
            wb.close()
            window.destroy()

        # 저장
        else:
            saveVal = []
            # Compare 실행 O
            if not compareValue:
                saveVal = all_values
            # Compare 실행 X
            elif compareValue:
                saveVal = compareValue

            ws = wb.active
            # 기존 데이터 삭제
            for row in range(2, ws.max_row):
                ws.delete_rows(3)

            # 정렬된 값으로 데이터 수정
            for i, s in enumerate(saveVal):
                ws.append([s[0], float(s[1]), float(s[2])])

            # 테두리
            for row in range(3, ws.max_row + 1):
                key = f'A{row}'
                ws[key].border = thin_border
                key = f'B{row}'
                ws[key].border = thin_border
                key = f'C{row}'
                ws[key].border = thin_border

            # 경북대학교 row 찾기
            for row in range(3, ws.max_row + 1):
                key = f'A{row}'
                if ws[key].value == '경북대학교':
                    print('\n경북대학교 :', row, ' line')
                    break
            key = f'A{row}'
            ws[key].fill = fill_back
            key = f'B{row}'
            ws[key].fill = fill_back
            key = f'C{row}'
            ws[key].fill = fill_back

            wb.save(filename)
            print('\nSave File [' + filename + ']\nCreate')

            wb.close()
            window.destroy()
    except:
        print('\nERROR')
        window.destroy()


btn_open = Button(window, text="Open Excel", width=20, command=OpenEvent)
btn_view = Button(window, text="Excel Sheet View", width=20, command=ViewEvent)
btn_compare = Button(window, text="Compare", width=20, command=CompareEvent)
btn_save = Button(window, text="Save", width=20, command=SaveEvent)

btn_open.pack()
btn_view.pack()
btn_compare.pack()
btn_save.pack()

window.mainloop()